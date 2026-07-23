"""Build a plate-template .xlsx for a kit: a pipeline-readable list + a linked plate grid.

Column A/B/C = TPositionId / SPositionBC / control_type (what make_ngsfilter reads, plain values;
control wells pre-filled). To the right, an 8x12 plate grid whose cells are formulas mirroring the
SPositionBC column, so filling the list updates the plate view. The file uploads back through the
job-submission "Upload Excel" path unchanged.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

LETTERS = "ABCDEFGH"
NUMS = list(range(1, 13))
GRID_COL0 = 5   # column E holds the plate row labels; F..Q hold columns 1..12

KIND_FILL = {
    "positive": "16A34A", "sequencing": "DC2626", "pcr": "F59E0B",
    "extraction": "7C3AED", "negative": "DC2626",
}


def _list_row(letter_idx: int, num: int) -> int:
    return 2 + letter_idx * 12 + (num - 1)   # rows 2..97, order A1..A12,B1..B12,...


def build_control_template_xlsx(kit) -> bytes:
    controls = {c.position.upper(): c for c in kit.controls if c.position}

    wb = Workbook()
    ws = wb.active
    ws.title = "plate"

    ws["A1"], ws["B1"], ws["C1"] = "Position", "Sample Name", "Control type"
    for cell in ("A1", "B1", "C1"):
        ws[cell].font = Font(bold=True)

    # list rows for all 96 wells; control wells pre-filled + coloured
    for li, letter in enumerate(LETTERS):
        for num in NUMS:
            well = f"{letter}{num}"
            r = _list_row(li, num)
            ws.cell(row=r, column=1, value=well)
            c = controls.get(well)
            if c:
                nc = ws.cell(row=r, column=2, value=c.name)
                ws.cell(row=r, column=3, value=c.kind.value)
                nc.fill = PatternFill("solid", fgColor=KIND_FILL.get(c.kind.value, "DC2626"))
                nc.font = Font(color="FFFFFF", bold=True)

    # plate grid header (well column numbers)
    for num in NUMS:
        hc = ws.cell(row=1, column=GRID_COL0 + num, value=num)
        hc.font = Font(bold=True)
        hc.alignment = Alignment(horizontal="center")

    # plate grid: each cell mirrors the list's SPositionBC via a formula
    for li, letter in enumerate(LETTERS):
        grow = 2 + li
        ws.cell(row=grow, column=GRID_COL0, value=letter).font = Font(bold=True)
        for num in NUMS:
            cell = ws.cell(row=grow, column=GRID_COL0 + num, value=f"=B{_list_row(li, num)}")
            cell.alignment = Alignment(horizontal="center")
            c = controls.get(f"{letter}{num}")
            if c:
                cell.fill = PatternFill("solid", fgColor=KIND_FILL.get(c.kind.value, "DC2626"))
                cell.font = Font(color="FFFFFF", bold=True)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 13
    for num in NUMS:
        ws.column_dimensions[get_column_letter(GRID_COL0 + num)].width = 13
    ws.cell(row=11, column=GRID_COL0, value=(
        "Fill sample names in the 'Sample Name' column (left); the plate grid mirrors them. "
        "Control rows are pre-filled — do not rename."))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
