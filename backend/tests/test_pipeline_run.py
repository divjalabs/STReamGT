import os

import pytest

from app.models.enums import ResultKind
from app.worker import pipeline_run as pr


def test_build_nextflow_cmd_has_no_outdir_override():
    cmd = pr.build_nextflow_cmd(
        pipeline_dir="/app/pipeline", input_tsv="/s/input.tsv", run_dir="/s",
        profile="docker", min_identity=0.9, min_overlap=20,
    )
    assert "run" in cmd and cmd[cmd.index("run") + 1].endswith("main.nf")
    assert "--outdir" not in cmd            # rely on pipeline default (kit_id nesting)
    assert "docker" in cmd and "--min_identity" in cmd
    assert "--expected_read_number" not in cmd  # omitted when not provided


def test_build_nextflow_cmd_passes_expected_reads():
    cmd = pr.build_nextflow_cmd(
        pipeline_dir="/app/pipeline", input_tsv="/s/input.tsv", run_dir="/s",
        profile="docker", min_identity=0.9, min_overlap=20, expected_read_number=5_000_000,
    )
    assert "--expected_read_number" in cmd and "5000000" in cmd


def test_collect_results_maps_kinds(tmp_path):
    kit = "DIVJA240"
    res = tmp_path / kit / "results"
    rep = tmp_path / kit / "reports"
    res.mkdir(parents=True)
    rep.mkdir(parents=True)
    (res / f"{kit}_genotypes.txt").write_text("x")
    (res / f"{kit}_positions.txt").write_text("x")
    (res / f"{kit}_frequency_of_sequences_by_marker.txt").write_text("x")
    (res / f"{kit}_consensus_genotypes.txt").write_text("x")
    (rep / f"{kit}_reads_summary.csv").write_text("x")
    (rep / f"{kit}_report.html").write_text("x")
    (rep / f"{kit}_consensus_report.html").write_text("x")

    found = pr.collect_results(str(tmp_path), kit)
    kinds = {r.kind for r in found}
    assert kinds == {
        ResultKind.genotypes, ResultKind.positions, ResultKind.frequency,
        ResultKind.consensus, ResultKind.reads_summary,
        ResultKind.html_report, ResultKind.consensus_report,
    }
    assert pr.find_result(found, ResultKind.genotypes).endswith("_genotypes.txt")
    # files whose suffix overlaps a shorter one must map to the more specific kind
    assert pr.find_result(found, ResultKind.consensus).endswith("_consensus_genotypes.txt")
    assert pr.find_result(found, ResultKind.genotypes).endswith(f"{kit}_genotypes.txt")
    assert pr.find_result(found, ResultKind.html_report).endswith(f"{kit}_report.html")
    assert pr.find_result(found, ResultKind.consensus_report).endswith("_consensus_report.html")


def test_collect_results_empty_when_missing(tmp_path):
    assert pr.collect_results(str(tmp_path), "NOPE") == []


def test_count_fastq_reads(tmp_path):
    import gzip
    p = tmp_path / "reads_1.fastq.gz"
    with gzip.open(p, "wt") as fh:
        for i in range(5):
            fh.write(f"@r{i}\nACGT\n+\nIIII\n")
    assert pr.count_fastq_reads(str(p)) == 5
    assert pr.count_fastq_reads(str(p), stop_at=3) == 3      # early-stop reports "at least 3"
    assert pr.count_fastq_reads(str(p), stop_at=100) == 5    # fewer than asked -> true count


def test_samples_text_to_rows_variants():
    rows = pr.samples_text_to_rows("TPositionId,SPositionBC\nA1,SAMP1\nB2,SAMP2\n")
    assert rows == [{"TPositionId": "A1", "SPositionBC": "SAMP1"},
                    {"TPositionId": "B2", "SPositionBC": "SAMP2"}]
    # tab and whitespace separated, header skipped
    assert pr.samples_text_to_rows("A1\tS1\nC3 S3")[0]["SPositionBC"] == "S1"


def test_samples_text_bad_line_raises():
    with pytest.raises(ValueError):
        pr.samples_text_to_rows("A1\nB2\n")  # only one field per line


def test_write_sample_xlsx_roundtrip(tmp_path):
    from openpyxl import load_workbook

    dest = tmp_path / "plate.xlsx"
    pr.write_sample_xlsx([{"TPositionId": "A1", "SPositionBC": "S1"}], str(dest))
    wb = load_workbook(dest)
    ws = wb.active
    assert [c.value for c in ws[1]] == ["TPositionId", "SPositionBC"]
    assert [c.value for c in ws[2]] == ["A1", "S1"]
