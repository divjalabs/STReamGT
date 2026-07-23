from tests.conftest import bearer, register, user_id, panel_id


def _register_kit(client, admin_token, assigned_ids, code="DIVJA240"):
    payload = {
        "kit_code": code,
        "panel_id": panel_id("UA_primers"),
        "selected_tags": ["PP1", "PP2", "PP3", "PP4"],
        "assigned_user_ids": assigned_ids,
    }
    return client.post("/api/kits", json=payload, headers=bearer(admin_token))


def test_admin_registers_kit_from_panel(client, catalog, admin_token):
    r = client.post(
        "/api/kits",
        json={
            "kit_code": "DIVJA240",
            "panel_id": panel_id("UA_primers"),
            "selected_tags": ["PP1", "PP2", "PP3", "PP4"],
        },
        headers=bearer(admin_token),
    )
    assert r.status_code == 201, r.text
    kit = r.json()
    assert kit["species"] == "brown bear"          # denormalized from the panel
    assert kit["status"] == "sent"                 # default
    assert [t["name"] for t in kit["tag_columns"]] == ["PP1", "PP2", "PP3", "PP4"]
    assert kit["controls"][0]["name_pattern"] == "blank"  # default control
    assert kit["panel"]["code"] == "UA_primers"


def test_position_controls_autoname_and_validate(client, catalog, admin_token):
    r = client.post(
        "/api/kits",
        json={
            "kit_code": "DIVJA300",
            "panel_id": panel_id("UA_primers"),
            "selected_tags": ["PP1", "PP2"],
            "controls": [
                {"kind": "sequencing", "name_pattern": "blank"},   # legacy pattern row
                {"kind": "pcr", "position": "a1"},                 # auto-named, lowercased well
                {"kind": "sequencing", "position": "b1"},          # token 'blank', not 'sequencing'
                {"kind": "extraction", "position": "c1"},          # token 'ext'
                {"kind": "positive", "position": "H12", "name": "myPos"},
            ],
        },
        headers=bearer(admin_token),
    )
    assert r.status_code == 201, r.text
    controls = r.json()["controls"]
    by_pos = {c["position"]: c for c in controls if c["position"]}
    assert by_pos["A1"]["name"] == "DIVJA300_pcr_A1"       # auto-generated
    assert by_pos["B1"]["name"] == "DIVJA300_blank_B1"     # sequencing -> 'blank' token
    assert by_pos["C1"]["name"] == "DIVJA300_ext_C1"       # extraction -> 'ext' token
    assert by_pos["H12"]["name"] == "myPos"                # explicit kept
    # KitSummary (list) also carries controls
    summary = [k for k in client.get("/api/kits", headers=bearer(admin_token)).json()
               if k["kit_code"] == "DIVJA300"][0]
    assert {c["kind"] for c in summary["controls"] if c["position"]} == \
        {"pcr", "sequencing", "extraction", "positive"}


def test_download_control_template(client, catalog, admin_token):
    import io
    from openpyxl import load_workbook
    r = client.post(
        "/api/kits",
        json={"kit_code": "DIVJA310", "panel_id": panel_id("UA_primers"), "selected_tags": ["PP1"],
              "controls": [{"kind": "pcr", "position": "B1"}]},
        headers=bearer(admin_token),
    )
    kit_id = r.json()["id"]
    resp = client.get(f"/api/kits/{kit_id}/control-template.xlsx", headers=bearer(admin_token))
    assert resp.status_code == 200
    assert "spreadsheet" in resp.headers["content-type"]
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert [ws.cell(1, c).value for c in (1, 2, 3)] == ["Position", "Sample Name", "Control type"]
    # the pcr control at B1 is pre-filled with its name + type
    filled = {ws.cell(row, 1).value: (ws.cell(row, 2).value, ws.cell(row, 3).value)
              for row in range(2, 98) if ws.cell(row, 2).value}
    assert filled["B1"] == ("DIVJA310_pcr_B1", "pcr")


def test_duplicate_control_well_rejected(client, catalog, admin_token):
    r = client.post(
        "/api/kits",
        json={
            "kit_code": "DIVJA301", "panel_id": panel_id("UA_primers"),
            "selected_tags": ["PP1"],
            "controls": [{"kind": "pcr", "position": "A1"}, {"kind": "positive", "position": "A1"}],
        },
        headers=bearer(admin_token),
    )
    assert r.status_code == 422 and "A1" in r.text


def test_control_templates_crud(client, catalog, admin_token, user_token):
    body = {"name": "std-layout", "positions": [
        {"kind": "pcr", "position": "A1"}, {"kind": "sequencing", "position": "B1"}]}
    r = client.post("/api/control-templates", json=body, headers=bearer(admin_token))
    assert r.status_code == 201, r.text
    tid = r.json()["id"]
    assert len(r.json()["positions"]) == 2
    # non-admin forbidden
    assert client.post("/api/control-templates", json=body, headers=bearer(user_token)).status_code == 403
    # duplicate name rejected
    assert client.post("/api/control-templates", json=body, headers=bearer(admin_token)).status_code == 409
    assert any(t["id"] == tid for t in client.get("/api/control-templates", headers=bearer(admin_token)).json())
    assert client.delete(f"/api/control-templates/{tid}", headers=bearer(admin_token)).status_code == 204


def test_unknown_tag_columns_rejected(client, catalog, admin_token):
    r = client.post(
        "/api/kits",
        json={"kit_code": "K1", "panel_id": panel_id(), "selected_tags": ["PP1", "PP99"]},
        headers=bearer(admin_token),
    )
    assert r.status_code == 422 and "PP99" in r.text


def test_kit_reads_crud(client, catalog, admin_token, user_token, monkeypatch):
    monkeypatch.setattr("app.api.kits.storage.object_exists", lambda key: True)
    tok_a = register(client, "a@x.com")
    tok_b = register(client, "b@x.com")
    kit_id = _register_kit(client, admin_token, [user_id("a@x.com")]).json()["id"]
    k1 = f"reads/kit/{kit_id}/u/reads_1.fastq.gz"
    k2 = f"reads/kit/{kit_id}/u/reads_2.fastq.gz"
    body = {"fastq1_key": k1, "fastq2_key": k2, "fastq1_name": "r1", "fastq2_name": "r2",
            "size1": 10, "size2": 20}

    # assigned user can register; outsider cannot
    assert client.put(f"/api/kits/{kit_id}/reads", json=body, headers=bearer(tok_b)).status_code == 403
    r = client.put(f"/api/kits/{kit_id}/reads", json=body, headers=bearer(tok_a))
    assert r.status_code == 200 and r.json()["fastq1_key"] == k1
    assert client.get(f"/api/kits/{kit_id}/reads", headers=bearer(tok_a)).json()["fastq2_name"] == "r2"

    # keys outside the kit's prefix are rejected
    bad = {**body, "fastq1_key": "uploads/9/x/evil.gz"}
    assert client.put(f"/api/kits/{kit_id}/reads", json=bad, headers=bearer(tok_a)).status_code == 422

    # delete clears it
    assert client.delete(f"/api/kits/{kit_id}/reads", headers=bearer(tok_a)).status_code == 204
    assert client.get(f"/api/kits/{kit_id}/reads", headers=bearer(tok_a)).json() is None


def test_access_control(client, catalog, admin_token):
    tok_a = register(client, "a@x.com")
    tok_b = register(client, "b@x.com")
    # kit assigned only to user A
    r = _register_kit(client, admin_token, [user_id("a@x.com")])
    assert r.status_code == 201
    kit_id = r.json()["id"]

    # A sees it; B does not; admin sees it
    assert [k["kit_code"] for k in client.get("/api/kits", headers=bearer(tok_a)).json()] == ["DIVJA240"]
    assert client.get("/api/kits", headers=bearer(tok_b)).json() == []
    assert len(client.get("/api/kits", headers=bearer(admin_token)).json()) == 1

    # direct get: A ok, B forbidden
    assert client.get(f"/api/kits/{kit_id}", headers=bearer(tok_a)).status_code == 200
    assert client.get(f"/api/kits/{kit_id}", headers=bearer(tok_b)).status_code == 403


def test_status_transitions(client, catalog, admin_token):
    tok_a = register(client, "a@x.com")
    kit_id = _register_kit(client, admin_token, [user_id("a@x.com")]).json()["id"]

    # client may set received, not analysed
    assert client.patch(f"/api/kits/{kit_id}", json={"status": "received"}, headers=bearer(tok_a)).json()["status"] == "received"
    assert client.patch(f"/api/kits/{kit_id}", json={"status": "analysed"}, headers=bearer(tok_a)).status_code == 403
    # client cannot self-approve a re-analysis either
    assert client.patch(f"/api/kits/{kit_id}", json={"status": "reanalyse"}, headers=bearer(tok_a)).status_code == 403
    # admin may set anything, including reanalyse
    assert client.patch(f"/api/kits/{kit_id}", json={"status": "analysed"}, headers=bearer(admin_token)).json()["status"] == "analysed"
    assert client.patch(f"/api/kits/{kit_id}", json={"status": "reanalyse"}, headers=bearer(admin_token)).json()["status"] == "reanalyse"


def test_non_admin_cannot_create_kit(client, catalog, user_token):
    assert _register_kit(client, user_token, []).status_code == 403


def test_delete_kit(client, catalog, admin_token):
    kit_id = _register_kit(client, admin_token, []).json()["id"]
    assert client.delete(f"/api/kits/{kit_id}", headers=bearer(admin_token)).status_code == 204
    assert client.get(f"/api/kits/{kit_id}", headers=bearer(admin_token)).status_code == 404
